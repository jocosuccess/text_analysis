import ntpath
import xlsxwriter
import openpyxl


class ControlFile:
    def __init__(self, path):
        self.path = path

    def extract_file_name(self):
        """
        Extract file name from file path
        :return: file name
        """

        path = ntpath.basename(self.path)
        head, tail = ntpath.split(path)
        filename = tail.replace(".txt", "")

        return filename

    def create_excel_file(self):
        """
        Create Excel file by given path
        :return: nothing
        """

        workbook = xlsxwriter.Workbook(self.path)
        worksheet = workbook.add_worksheet()

        fields = (['No', 'FILENAME', 'FOG_INDEX', 'KINCAID_INDEX', 'NUMWORD', 'NEGATIVE', 'POSITIVE',
                   'UNCERTAINTY', 'LITIGIOUS', 'STRONG MODAL', 'WEAK MODAL', 'CONSTRAINING'])

        row = 0
        col = 0

        for item in fields:

            worksheet.write(row, col, item)
            col += 1

        workbook.close()

    def write_in_excel(self, content):
        """
        Write content into given excel file
        :param content: content to be written
        :return: nothing
        """

        oxl = openpyxl.load_workbook(self.path)
        sheet = oxl.get_sheet_by_name('Sheet1')
        sheet.append(content)

        oxl.save(self.path)
        oxl.close()

    def read_loughran_words(self):
        """
        Read loughran words from loughran excel file
        :return: loughran words as dict
        """

        wb = openpyxl.load_workbook(self.path)
        init_sheet_names = wb.get_sheet_names()
        origin_sheet_names = init_sheet_names[1: len(init_sheet_names)]
        sheet_names = [name.upper() for name in origin_sheet_names]
        loughran_words = {}

        for i, sheet_name in enumerate(sheet_names):

            sheet = wb.get_sheet_by_name(origin_sheet_names[i])

            for row in range(1, sheet.max_row + 1):
                # Each row in the spreadsheet has data for one census tract.
                word = sheet['A' + str(row)].value

                if sheet_name not in loughran_words.keys():
                    loughran_words[sheet_name] = [word]
                else:
                    loughran_words[sheet_name].append(word)

        return loughran_words
